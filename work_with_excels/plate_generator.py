# -*- coding: utf-8 -*-
import time
import copy
import datetime
import openpyxl

title = [u"序号", u"自选号码", u"号段时间"]

def handleStr(string):
    return string.strip().replace(',','').replace('.','').replace(' ','').replace('\n','').replace('\t','')

def isThreeSame(string):
    flag = -1
    for num in string:
        if num.isdigit():
            flag = num
            break

    c = 0
    for num in string:
        if num == flag:
            c += 1

    return c >= 3

if __name__ == "__main__":
    w_excel = openpyxl.Workbook()
    w_sheet = w_excel.active
    for i in range(len(title)):
        w_sheet.cell(row = 1, column = i + 1, value = title[i])
    excel_row_index = 2

    r_excel = openpyxl.load_workbook("haoduan.xlsx")
    sheet = r_excel.active
    for row_index in range(2, sheet.max_row + 1):
        origin_number = handleStr(str(sheet["A" + str(row_index)].value))
        output_time = sheet["B" + str(row_index)].value

        [start, end] = origin_number.split('~')
        self_construct = []
        for i in range(len(start)):
            if start[i] != end[i]:
                self_construct.append(i)
        assert len(self_construct) == 2

        if int(start[2]) in [6, 8, 1, 9, 2]:
            for item in ["68", "88", "66", "86", "99", "98"]:
                new_number = list(start)
                new_number[3] = item[0]
                new_number[5] = item[1]
                new_number = ''.join(new_number)
                if ("A" in new_number and "V" in new_number) or isThreeSame(new_number) or "S" in new_number or "X" in new_number or "D" in new_number or "P" in new_number or "J" in new_number or "F" in new_number or "Q" in new_number or ("W" in new_number and "C" in new_number) or ("W" in new_number and "R" in new_number) or ("W" in new_number and "K" in new_number) or "W2" in new_number or "D8" in new_number or "B" in new_number:
                    continue
                w_sheet.cell(row = excel_row_index, column = 1, value = str(excel_row_index - 1))
                w_sheet.cell(row = excel_row_index, column = 2, value = new_number)
                w_sheet.cell(row = excel_row_index, column = 3, value = output_time)
                excel_row_index += 1

    xlsname = "自选结果.xlsx"
    w_excel.save(xlsname)
    print(u"生成文件 %s" % xlsname)
